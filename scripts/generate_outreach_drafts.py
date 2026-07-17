# Generate honest, value-first outreach drafts for the 17-company formal list.
#
# PRINCIPLES (from project memory — violating these poisons the whole batch):
#   - NO fabricated cases, numbers or clients. Machine specs only from the DB
#     (JL-L-2TZP600: three-side seal / doypack / zipper, up to 220 pcs/min).
#   - Value-first: the email shares two published guides; the pitch is one soft line.
#   - Every email carries a real signature, the company address, and an opt-out line.
#   - The old drafts/ folder (2026-07-02) contained a fabricated 45-to-12-minute
#     changeover story — those drafts are superseded by this batch. Do not reuse them.
#
# Output: C:/dev/marketing strategy/reylong-leads/output/drafts-guides-202607/
#   one .txt per company + send_manifest.csv (email channel only).
import csv
import io
import re
import sys
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE = Path(r"C:/dev/marketing strategy/reylong-leads/output")
# 2026-07-16 起名單唯一來源是 開發名單_MASTER.xlsx（正式開發名單分頁）；
# 舊的 formal_outreach_leads_consolidated.csv 已整併刪除。
SRC = BASE / "開發名單_MASTER.xlsx"
OUT = BASE / "drafts-guides-202607"
OUT.mkdir(exist_ok=True)

GUIDE_HEATSEAL = "https://www.reylong.com/news/heat-seal-strength-failure-diagnosis"
GUIDE_AI = "https://www.reylong.com/news/manual-vs-ai-inspection-woven-bag-lines"

SIGNATURE = """Best regards,

William Ho
Rey Long Machinery Industrial Co., Ltd.
No. 3, Zhuwei Rd., Douliu City, Yunlin County, Taiwan
https://www.reylong.com

P.S. If you'd rather not hear from us again, just reply "unsubscribe" — we won't write twice."""

# One personal line per company, grounded ONLY in the verified company_background
# from the consolidated CSV. No invented facts.
PERSONAL = {
    "American Packaging Corporation": "your flexible packaging and pouch operations",
    "C-P Flexible Packaging": "your custom packaging and rollstock lines",
    "Poly Print": "your printing and laminating operations",
    "Printpack": "your flexible packaging lines across food and medical markets",
    "Bison Bag Co.": "your stand-up and spouted pouch lines",
    "CarePac": "your mylar pouch production for food and cosmetics",
    "Eagle Flexible Packaging": "your custom printed pouch and rollstock lines",
    "Eastern Web Handling": "your custom stand-up pouch manufacturing",
    "Glenroy, Inc.": "your stand-up, zipper and spouted pouch lines",
    "IMPAK Corporation": "your in-house bag making for ZipSeal and flat pouches",
    "LPS Industries": "your custom pouch and barrier bag production",
    "Pregis Sharp Systems": "your poly bag and bagging system operations",
    "SunDance USA": "your digital flexible packaging production",
    "TricorBraun Flex": "your pouch converting operations",
    "ePac Flexible Packaging": "your short-run pouch making sites",
}

EN_BODY = """Subject: A heat-seal diagnostic guide your production team may find useful

Hi {company} team,

Rey Long builds bag- and pouch-making machinery in Taiwan. We recently started publishing the troubleshooting guides we wished existed when our own customers call us with production problems — no registration wall, no sales pitch inside:

* Heat-Seal Strength Failures: the four distinct failure modes (false seal, burn-through, channel leak, random weak seal) and the order to work through temperature, dwell, pressure and cooling —
  {g1}

* Manual vs AI Inspection: an honest comparison built on published research, including the four cases where manual inspection is still the right call —
  {g2}

Both apply to whatever equipment you run today. And if {personal} ever need a second source for pouch machines — three-side seal, doypack and zipper formats on one platform, up to 220 pcs/min — we would be glad to talk.

{signature}
"""

# Japanese draft for the two JP companies (contact-form channel, pasted manually).
JA_BODY = """件名: ヒートシール不良の診断ガイドのご案内（台湾・製袋機メーカー Rey Long）

{company} ご担当者様

突然のご連絡失礼いたします。台湾で製袋機械を製造しております Rey Long Machinery の William Ho と申します。

このたび、現場の生産技術者様向けに、実務的な技術ガイドを公開いたしました。登録不要・営業資料ではございません。

・ヒートシール強度不良の診断ガイド（4つの不良モードの見分け方と、温度・加圧時間・圧力・冷却の確認順序）
  {g1}

・目視検査とAI検査の比較（公開研究データに基づく比較。目視検査が適する4つのケースも正直に記載）
  {g2}

いずれも、現在お使いの設備を問わずご活用いただける内容です。もし三方袋・スタンドパック・チャック付き袋の製袋機（1台で最大220袋/分）について、セカンドソースをご検討の機会がございましたら、お気軽にお声がけください。

何卒よろしくお願い申し上げます。

Rey Long Machinery Industrial Co., Ltd.
William Ho
台湾雲林県斗六市竹囲路3号
https://www.reylong.com

※今後のご案内が不要な場合は、その旨ご返信いただければ以後お送りいたしません。
"""

wb = openpyxl.load_workbook(SRC, read_only=True)
ws = wb["正式開發名單"]
headers = [str(c.value or "") for c in ws[1]]
rows = []
for raw in ws.iter_rows(min_row=2, values_only=True):
    if not raw or not raw[0]:
        continue
    d = dict(zip(headers, (str(v) if v is not None else "" for v in raw)))
    rows.append({
        "company_name": d.get("公司名稱", ""),
        "country": d.get("國家", ""),
        "contact_email": d.get("Email", ""),
        "outreach_channel": d.get("寄送通道", ""),
        "contact_page_or_form": d.get("聯絡頁或表單", ""),
    })
wb.close()

manifest = []
email_re = re.compile(r"[\w.+-]+@[\w-]+\.[\w.]+")

for row in rows:
    company = row["company_name"].strip()
    country = row["country"].strip()
    emails = email_re.findall(row.get("contact_email") or "")
    channel = row.get("outreach_channel", "").strip()

    if country == "JP":
        text = JA_BODY.format(company=company, g1=GUIDE_HEATSEAL, g2=GUIDE_AI)
    else:
        personal = PERSONAL.get(company, "your production lines")
        text = EN_BODY.format(
            company=company.replace(", Inc.", "").replace(" LLC", "").replace(" Co.", ""),
            personal=personal, g1=GUIDE_HEATSEAL, g2=GUIDE_AI, signature=SIGNATURE,
        )

    safe = re.sub(r"[^\w\- ]", "", company)[:50].strip()
    (OUT / f"{safe}.txt").write_text(text, encoding="utf-8")

    # Only the first email per company goes into the send manifest; extra addresses
    # (e.g. Pregis's seven inboxes) stay in the CSV for manual judgement.
    manifest.append({
        "company": company,
        "country": country,
        "channel": "email" if emails else ("form" if "form" in channel.lower() or row.get("contact_page_or_form") else "none"),
        "to": emails[0] if emails else "",
        "draft_file": f"{safe}.txt",
        "contact_page": (row.get("contact_page_or_form") or "").split(";")[0].strip(),
    })

with open(OUT / "send_manifest.csv", "w", newline="", encoding="utf-8-sig") as f:
    w = csv.DictWriter(f, fieldnames=["company", "country", "channel", "to", "draft_file", "contact_page"])
    w.writeheader()
    w.writerows(manifest)

email_n = sum(1 for m in manifest if m["channel"] == "email")
form_n = sum(1 for m in manifest if m["channel"] == "form")
print(f"drafts written: {len(manifest)} (email: {email_n}, form: {form_n}, none: {len(manifest)-email_n-form_n})")
print(f"output: {OUT}")
