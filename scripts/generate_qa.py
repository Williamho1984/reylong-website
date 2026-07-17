import openpyxl, sys, io
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

# ── Q&A Dataset ───────────────────────────────────────────────────────────────
# Format: (category, question_en, answer_en, related_product, priority)
# priority: H=high (must answer), M=medium, L=lower

QA = [

    # ── CATEGORY 1: Product Selection / Which Machine? ──────────────────────
    ('Product Selection',
     'Which machine should I choose for making stand-up doypack pouches?',
     'The JL-L-2TZP600 is the right choice. It produces stand-up doypack pouches — both plain and with zipper — at 35–120 pcs/min, using heat-seal laminated film (30–180 μm). It handles five bag styles on a single platform so you can switch formats without changing machines.',
     'JL-L-2TZP600', 'H'),

    ('Product Selection',
     'Can one machine make both flat three-side seal bags and doypack pouches?',
     'Yes. The JL-L-2TZP600 is a multi-function platform that produces three-side seal bags, three-side seal zipper bags, four-side seal bags, doypack pouches, and doypack with zipper — all on one machine. Switching between bag styles requires a format changeover but no separate machine.',
     'JL-L-2TZP600', 'H'),

    ('Product Selection',
     'I need to produce coffee bags with a degassing valve. Which machine is suitable?',
     'The JL-L-2TZP600 supports coffee packaging (both center-seal and doypack formats) and can accommodate degassing valve insertion. Please contact us with your specific valve type and bag dimensions for a configuration confirmation.',
     'JL-L-2TZP600', 'H'),

    ('Product Selection',
     'Do you have a machine for PP woven bags?',
     'Yes. The JLPTCSM-1300W is a fully integrated convention line for high-volume PP woven bag production, combining flexographic printing, tube forming, cutting & sewing, and overtape application in one continuous process at 25–40 bags/min.',
     'JLPTCSM-1300W', 'H'),

    ('Product Selection',
     'What machine makes medical packaging bags or sterilization pouches?',
     'The JL-L-2TZP600 supports medical packaging applications, including individual bags and flat pouches. For specialized medical bag types such as gusseted tube bags or four-layer header bags, please contact our engineering team with your specifications.',
     'JL-L-2TZP600', 'H'),

    ('Product Selection',
     'I need to package pet food. What bag machine do you recommend?',
     'The JL-L-2TZP600 handles pet food packaging in doypack, stand-up zipper, and three-side seal formats using high-barrier laminated film. This is a common application — please share your target bag dimensions and production speed for a detailed recommendation.',
     'JL-L-2TZP600', 'H'),

    ('Product Selection',
     'Which machine is best for facial mask or cosmetic packaging?',
     'The JL-L-2TZP600 produces three-side seal and doypack bags commonly used for facial masks, skincare, and cosmetic products. It supports printed laminated films with ≤0.3 mm positional accuracy for precise register printing alignment.',
     'JL-L-2TZP600', 'M'),

    ('Product Selection',
     'Can your machines handle anti-static packaging for electronics?',
     'Yes. The JL-L-2TZP600 can process anti-static laminated films suitable for electronic component packaging. The applicable film thickness range is 30–180 μm. Please specify your film type so we can confirm compatibility.',
     'JL-L-2TZP600', 'M'),

    # ── CATEGORY 2: JL-L-2TZP600 Technical Specs ────────────────────────────
    ('Technical – JL-L-2TZP600',
     'What is the maximum production speed of the JL-L-2TZP600?',
     'The JL-L-2TZP600 runs at up to 220 pcs/min for three-side seal bags, 150 pcs/min for three-side seal with zipper bags, and 120 pcs/min for doypack pouches. Minimum speed is 35 pcs/min across all formats.',
     'JL-L-2TZP600', 'H'),

    ('Technical – JL-L-2TZP600',
     'What bag sizes can the JL-L-2TZP600 produce?',
     'Bag width ranges from 100 to 580 mm, and bag length from 50 to 420 mm. A skip function extends the effective length up to 6× for longer bags. Bottom seal width is adjustable from 5 to 60 mm.',
     'JL-L-2TZP600', 'H'),

    ('Technical – JL-L-2TZP600',
     'What materials can the JL-L-2TZP600 process?',
     'The machine is designed for heat-seal laminated films with a thickness of 30–180 μm. Common structures include NY/PE, PET/PE, AL/PE, OPP/CPP, and BOPP/LLDPE. Please confirm your specific film structure for compatibility.',
     'JL-L-2TZP600', 'H'),

    ('Technical – JL-L-2TZP600',
     'What power supply does the JL-L-2TZP600 require?',
     '220 V / 3-phase / 60 Hz, with a total power consumption of 60 kW. If your facility operates at 50 Hz, please inform us — the machine can be configured for 50 Hz with appropriate servo adjustments.',
     'JL-L-2TZP600', 'H'),

    ('Technical – JL-L-2TZP600',
     'How precise is the sealing and registration on the JL-L-2TZP600?',
     'Positional accuracy is ≤0.3 mm, achieved through full servo-motor control (Panasonic). The sealing system uses 4 groups of heating + 2 groups of cooling for both vertical and horizontal seals, ensuring consistent seal quality.',
     'JL-L-2TZP600', 'H'),

    ('Technical – JL-L-2TZP600',
     'How is the zipper sealed on the zipper bag format?',
     'Zipper sealing uses ultrasonic technology (1 group), which provides a clean, consistent seal without heat damage to the zipper track. Standard zipper width is 13 mm.',
     'JL-L-2TZP600', 'M'),

    ('Technical – JL-L-2TZP600',
     'What is the machine weight and footprint of the JL-L-2TZP600?',
     'The machine weighs 9,000 kg. Please contact us for the full installation drawing with exact footprint dimensions and recommended clearance space for your facility planning.',
     'JL-L-2TZP600', 'M'),

    ('Technical – JL-L-2TZP600',
     'What control system does the JL-L-2TZP600 use?',
     'The machine uses a PLC touch-screen interface with full servo-motor control (Panasonic servo drives). The 10 mm steel-plate frame provides vibration resistance for high-speed continuous operation.',
     'JL-L-2TZP600', 'M'),

    ('Technical – JL-L-2TZP600',
     'Can the JL-L-2TZP600 run retort or high-temperature packaging films?',
     'Yes. The temperature range reaches up to 300°C, making it suitable for retort packaging films (NY/CPP, AL/CPP retort-grade) used in cooked food and sterilized product packaging.',
     'JL-L-2TZP600', 'M'),

    # ── CATEGORY 3: PP Woven Bag Line (JLPTCSM-1300W) ───────────────────────
    ('Technical – JLPTCSM-1300W',
     'What is the production speed of the PP woven bag convention line?',
     'The JLPTCSM-1300W produces 25–40 bags/min. It integrates four stages — flexographic printing, tube forming, cutting & sewing, and overtape application — in a single continuous process.',
     'JLPTCSM-1300W', 'H'),

    ('Technical – JLPTCSM-1300W',
     'What bag dimensions does the JLPTCSM-1300W handle?',
     'The line handles fabric widths of 700–1300 mm (before tubing), resulting in finished bag widths of 430–660 mm. Cutting length ranges from 550–1250 mm. Gusset depth is adjustable from 40–90 mm.',
     'JLPTCSM-1300W', 'M'),

    ('Technical – JLPTCSM-1300W',
     'How many print colors does the woven bag convention line support?',
     'The JLPTCSM-1300W supports 4-color flexographic printing (4+0 configuration) with a printing repeat length of 450–1200 mm and max printing width of 1300 mm.',
     'JLPTCSM-1300W', 'M'),

    # ── CATEGORY 4: Flexographic Printing Machine ───────────────────────────
    ('Technical – Printing Machine',
     'What materials can the 6-color flexographic printing machine print on?',
     'The JLRPM-6800BO/6C is designed for PP woven fabric and PP laminated fabric, with a max web width of 800 mm and max machine speed of 120 m/min (printing speed up to 100 m/min).',
     'JLRPM-6800BO/6C', 'M'),

    ('Technical – Printing Machine',
     'How many colors can the flexographic printing machine print?',
     '6 colors, using 6 independent printing stations with ceramic anilox rollers. Configurations supported: 0+6, 1+5, 2+4, 3+3 (front/back). An inter-color drying system ensures clean separation between stations.',
     'JLRPM-6800BO/6C', 'M'),

    # ── CATEGORY 5: Eddy Current Separator ──────────────────────────────────
    ('Technical – Recycling',
     'What materials does the eddy current separator recover?',
     'The JLECS-1000W recovers non-ferrous metals (aluminum, copper, brass) from mixed material streams such as e-waste, shredded plastic, and mixed scrap. It combines magnetic drum separation (ferrous removal) with eddy current separation (non-ferrous recovery) at 1,000 kg/h capacity.',
     'JLECS-1000W', 'M'),

    # ── CATEGORY 6: AI / Smart Machine ──────────────────────────────────────
    ('AI Solutions',
     'What is the AI Machine Intelligence solution?',
     'Rey Long\'s AI Machine Intelligence embeds edge computing and industrial IoT into existing production lines. It provides real-time decisions at the machine level — including computer vision quality inspection, predictive maintenance, and process optimization — without relying on cloud connectivity.',
     'AI Solutions', 'M'),

    ('AI Solutions',
     'Can AI be retrofitted onto an existing machine?',
     'Yes. The AI program is designed case-by-case for specific machines and processes. It can be integrated into existing production lines without requiring a full machine replacement. Contact us with your machine type and target application for an assessment.',
     'AI Solutions', 'M'),

    # ── CATEGORY 7: Procurement / Business ──────────────────────────────────
    ('Procurement',
     'How do I get a price quote?',
     'Pricing depends on your specific configuration — bag type, dimensions, speed, and options. Use the "Request a Quote" button on any product page or visit our Contact page. Our team responds within 1 business day with a tailored quotation.',
     'All', 'H'),

    ('Procurement',
     'How much does the JL-L-2TZP600 cost?',
     'Machine pricing varies based on configuration, options, and destination. To receive an accurate quote for your requirements, please use the Request a Quote form on the product page — our team will respond within 1 business day.',
     'JL-L-2TZP600', 'H'),

    ('Procurement',
     'What is the price of the PP woven bag convention line?',
     'Pricing for the JLPTCSM-1300W depends on your production specifications. Please fill in the inquiry form on the product page and our sales team will provide a detailed quotation promptly.',
     'JLPTCSM-1300W', 'H'),

    ('Procurement',
     'Can you give me a ballpark price or price range?',
     'We quote based on your exact requirements to ensure the price reflects your configuration accurately. Please submit an inquiry with your target bag type, dimensions, and speed — we\'ll get back to you within 1 business day with a specific figure.',
     'All', 'H'),

    ('Procurement',
     'What is the typical lead time for machine delivery?',
     'Standard lead time is 2–3 months from order confirmation. This covers manufacturing, quality testing, and pre-shipment inspection. Complex customizations may require additional time — we will confirm the exact schedule at order stage.',
     'All', 'H'),

    ('Procurement',
     'Do you offer machine customization?',
     'Yes. Rey Long offers customization for most machine models — including bag size range, speed configuration, optional accessories, and special film compatibility. Please provide your production requirements and our engineering team will advise on feasibility.',
     'All', 'H'),

    ('Procurement',
     'What payment terms do you offer?',
     'Payment terms are discussed on a per-order basis. Please submit an inquiry and our sales team will include payment options in the quotation response.',
     'All', 'H'),

    ('Procurement',
     'Do you export to the United States / Japan?',
     'Yes. Rey Long supplies machinery to customers in the United States, Japan, and globally. We have experience with export documentation, container loading, and international shipping coordination.',
     'All', 'H'),

    ('Procurement',
     'What voltage and frequency options are available?',
     'Machines are configurable for different power standards. The JL-L-2TZP600 default is 220V/3P/60Hz. 50 Hz configurations are available — please specify your local power standard when requesting a quote.',
     'All', 'H'),

    # ── CATEGORY 8: After-Sales / Support ───────────────────────────────────
    ('After-Sales',
     'What warranty does Rey Long provide?',
     'Rey Long provides warranty coverage on parts and workmanship. Specific warranty terms are stated in the sales contract. Please contact us for details applicable to your machine and region.',
     'All', 'H'),

    ('After-Sales',
     'Is installation and training included?',
     'Yes. Rey Long provides technical documentation, remote support, and on-site installation assistance. Operator training is included as part of the commissioning process.',
     'All', 'H'),

    ('After-Sales',
     'How do I order spare parts?',
     'Spare parts are available globally with fast delivery. Contact our after-sales team with your machine model number and the part specification (or a photo/drawing) for a quick quotation.',
     'All', 'H'),

    ('After-Sales',
     'Do you provide remote technical support?',
     'Yes. Rey Long provides remote support via video call, with real-time guidance from our engineers. For major issues, on-site service can be arranged.',
     'All', 'M'),

    # ── CATEGORY 9: Company Background ──────────────────────────────────────
    ('Company',
     'Where is Rey Long based?',
     'Rey Long Machinery Co., Ltd. is a professional manufacturer of plastic and packaging machinery headquartered in Taiwan, serving manufacturers worldwide.',
     'Company', 'H'),

    ('Company',
     'What types of machines does Rey Long manufacture?',
     'Rey Long manufactures flexible packaging bag making machines (three-side seal, doypack, zipper), PP woven bag convention lines, flexographic printing machines, recycling separation systems, and AI-powered machine intelligence solutions.',
     'Company', 'H'),

    ('Company',
     'Does Rey Long have CE certification or international certifications?',
     'CE certification for our machines is currently in progress. We will provide updated certification documentation as it becomes available. Please contact us for the latest status relevant to your target market.',
     'Company', 'M'),

    # ── CATEGORY 10: Application / Use-Case ─────────────────────────────────
    ('Application',
     'Can the JL-L-2TZP600 produce liquid packaging pouches?',
     'Yes. The JL-L-2TZP600 supports liquid packaging in doypack and three-side seal formats using appropriate barrier laminated films. For liquid-fill integration, the bag is typically made on the JL-L-2TZP600 and filled on a separate filling machine.',
     'JL-L-2TZP600', 'M'),

    ('Application',
     'Is the machine suitable for rice or grain packaging?',
     'Yes. The JL-L-2TZP600 handles rice packaging in three-side seal and doypack formats. Film thickness up to 180 μm accommodates the heavier-gauge films often used for grain packaging.',
     'JL-L-2TZP600', 'M'),

    ('Application',
     'Can I use the machine for garment or clothing packaging?',
     'Yes. Three-side seal bags and doypack pouches produced by the JL-L-2TZP600 are used for garment packaging. The machine supports wide bags (up to 580 mm) suitable for folded clothing.',
     'JL-L-2TZP600', 'M'),

    ('Application',
     'What frozen food packaging can this machine produce?',
     'The JL-L-2TZP600 produces frozen food packaging in three-side seal, doypack, and zipper formats using NY/PE or PET/PE laminated films rated for freezer temperatures. The sealing temperature range (up to 300°C) handles the full range of freeze-grade films.',
     'JL-L-2TZP600', 'M'),

    # ── CATEGORY 11: Comparison / Competitors ────────────────────────────────
    ('Comparison',
     'How does the JL-L-2TZP600 compare to other high-speed pouch machines on the market?',
     'The JL-L-2TZP600 delivers ≤0.3 mm positional accuracy and up to 220 pcs/min using Panasonic full-servo control — performance comparable to premium international machines, with competitive pricing and a 2–3 month lead time. The 10 mm steel-plate frame provides superior rigidity at high speed.',
     'JL-L-2TZP600', 'M'),

    ('Comparison',
     'What makes Rey Long machines different from other manufacturers?',
     'Key differentiators: 10 mm steel-plate frame (heavier than industry standard) for vibration resistance, full Panasonic servo control, multi-format capability on a single platform, AI integration options, and direct factory engineering support. Lead time is 2–3 months.',
     'All', 'M'),
    # ── CATEGORY 12: Technical Operations / Setup ────────────────────────────
    ('Technical – Operations',
     'How long does it take to change over between bag formats on the JL-L-2TZP600?',
     'Format changeover time depends on the degree of size change. Minor adjustments (same format, different size) typically take 15–30 minutes. Switching between bag types (e.g., three-side seal to doypack) may take 30–60 minutes with trained operators. Rey Long provides changeover procedures and operator training during commissioning.',
     'JL-L-2TZP600', 'H'),

    ('Technical – Operations',
     'How do I set the sealing temperature for a new film material?',
     'Sealing temperature is set via the PLC touch screen. Start with the film supplier\'s recommended heat-seal temperature range, then run test bags at low speed and adjust in 5°C increments until seal quality is consistent. The machine supports up to 300°C. Rey Long\'s commissioning engineer will demonstrate this process during installation.',
     'JL-L-2TZP600', 'H'),

    ('Technical – Operations',
     'What causes poor seal quality or seal leakage?',
     'Common causes: (1) sealing temperature too low for the film material, (2) sealing pressure insufficient, (3) sealing dwell time too short at high speed, (4) film contamination or moisture on the seal area, (5) worn sealing bar. Check temperature settings first, then inspect the sealing bar surface. Contact our technical support team if the issue persists.',
     'JL-L-2TZP600', 'H'),

    ('Technical – Operations',
     'What causes misaligned printing registration on finished bags?',
     'Registration errors are typically caused by: (1) inconsistent film tension — check the tension control setting, (2) film splice or quality variation, (3) photo-eye sensor sensitivity needs adjustment, (4) servo parameters need recalibration. The JL-L-2TZP600 has ≤0.3 mm accuracy under normal operating conditions. Contact Rey Long technical support for sensor and servo calibration guidance.',
     'JL-L-2TZP600', 'H'),

    ('Technical – Operations',
     'How do I thread the film into the machine?',
     'Film threading procedure is detailed in the operation manual provided with the machine. Rey Long\'s commissioning engineer demonstrates threading during installation and training. A threading diagram is also posted near the film path on the machine for reference.',
     'JL-L-2TZP600', 'M'),

    ('Technical – Operations',
     'What is the recommended maintenance schedule for the JL-L-2TZP600?',
     'Recommended schedule: Daily — clean sealing bars, check film path and tension rollers. Weekly — lubricate moving parts per the lubrication chart, inspect servo belts. Monthly — check electrical connections, inspect heating elements and thermocouples. A full maintenance checklist is included in the operation manual. Rey Long also offers remote maintenance support.',
     'JL-L-2TZP600', 'H'),

    ('Technical – Operations',
     'What spare parts should I keep in stock?',
     'Recommended stock items: sealing bars (vertical and horizontal), silicone pads, heating elements, thermocouples, photo-eye sensors, and servo belts. Rey Long provides a recommended spare parts list with each machine. Parts can be ordered directly from our after-sales team.',
     'JL-L-2TZP600', 'H'),

    ('Technical – Operations',
     'The machine stops mid-run and shows an error. What should I do?',
     'Check the PLC touch screen for the error code and alarm message. The operation manual contains a full error code reference. Common causes: film break (check tension and splice), servo fault (check drive indicator lights), temperature alarm (verify heating element and thermocouple). If you cannot resolve the issue, contact Rey Long technical support with the error code and a description of the situation.',
     'JL-L-2TZP600', 'H'),

    ('Technical – Operations',
     'Can I run the machine at maximum speed immediately after startup?',
     'No. It is recommended to warm up the sealing bars for 10–15 minutes before running at full speed, allowing the temperature to stabilize uniformly across all heating groups. Start at low speed, verify seal quality, then gradually increase to target speed.',
     'JL-L-2TZP600', 'M'),

    ('Technical – Operations',
     'How do I adjust the zipper insertion position on zipper bags?',
     'Zipper position is adjusted via the servo-controlled zipper feeding mechanism on the PLC interface. Set the target zipper-to-top-seal distance, then run a test bag and measure. Fine-tune in small increments. The commissioning engineer covers zipper setup in detail during training.',
     'JL-L-2TZP600', 'M'),

    ('Technical – Operations',
     'What film roll width and core diameter does the machine accept?',
     'Max web roll diameter is φ600 mm. The machine accepts standard 3-inch cores. Max web width for the main film is 1240 mm; patch web is 150 mm. Please confirm your roll specifications when ordering to ensure the unwind stand is configured correctly.',
     'JL-L-2TZP600', 'M'),

    ('Technical – Operations',
     'Can the machine handle pre-printed films with registration marks?',
     'Yes. The JL-L-2TZP600 uses photo-eye registration sensors to track print marks on pre-printed films and maintain ≤0.3 mm positional accuracy. Ensure your film has a consistent, high-contrast registration mark and inform us of the mark position when configuring the machine.',
     'JL-L-2TZP600', 'H'),

    ('Technical – Operations',
     'What happens if the film breaks during production?',
     'The machine automatically stops and triggers a film break alarm. Resplice the film, re-thread through the film path, reset the alarm on the PLC, and restart at low speed to verify the splice seals correctly before returning to full production speed.',
     'JL-L-2TZP600', 'M'),

    ('Technical – Operations',
     'Is it possible to produce bags longer than 420 mm?',
     'Yes. The skip function allows bag lengths beyond the standard 420 mm limit, extending effective length up to 6× the base pitch. Contact our engineering team with your target dimensions for configuration confirmation.',
     'JL-L-2TZP600', 'M'),

]

# ── Build Excel ────────────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = 'Q&A Knowledge Base'

# Header
headers = ['#', 'Category', 'Question (EN)', 'Answer (EN)', 'Related Product', 'Priority', 'Status', 'Notes']
header_fill = PatternFill('solid', fgColor='1F4E79')
header_font = Font(bold=True, color='FFFFFF', size=11)
for ci, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=ci, value=h)
    c.fill = header_fill
    c.font = header_font
    c.alignment = Alignment(horizontal='center', vertical='center')
ws.row_dimensions[1].height = 20

# Category colors
cat_colors = {
    'Product Selection':          'E2EFDA',  # green
    'Technical – JL-L-2TZP600':  'DDEBF7',  # blue
    'Technical – JLPTCSM-1300W':  'EAF2FF',  # light blue
    'Technical – Printing Machine':'FFF2CC', # yellow
    'Technical – Recycling':      'FCE4D6',  # orange
    'Technical – Operations':     'D6EAF8',  # steel blue
    'AI Solutions':               'F4CCFF',  # purple
    'Procurement':                'FFE4E1',  # pink
    'After-Sales':                'E8F5E9',  # mint
    'Company':                    'FFF9C4',  # pale yellow
    'Application':                'E0F7FA',  # cyan
    'Comparison':                 'F3E5F5',  # lavender
}
priority_colors = {'H': 'C6EFCE', 'M': 'FFEB9C', 'L': 'FFCCCC'}

for ri, (cat, q, a, prod, prio) in enumerate(QA, 2):
    row_data = [ri-1, cat, q, a, prod, prio, 'Draft', '']
    fill_color = cat_colors.get(cat, 'FFFFFF')
    for ci, val in enumerate(row_data, 1):
        c = ws.cell(row=ri, column=ci, value=val)
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if ci == 6:  # priority column
            c.fill = PatternFill('solid', fgColor=priority_colors.get(prio, 'FFFFFF'))
        else:
            c.fill = PatternFill('solid', fgColor=fill_color)
    ws.row_dimensions[ri].height = 60

# Column widths
widths = [4, 22, 50, 80, 22, 10, 10, 25]
for ci, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w

ws.freeze_panes = 'A2'

# Stats sheet
ws2 = wb.create_sheet('Summary')
cats = {}
for cat, q, a, prod, prio in QA:
    cats[cat] = cats.get(cat, 0) + 1
ws2.cell(1,1,'Category').font = Font(bold=True)
ws2.cell(1,2,'Count').font = Font(bold=True)
ws2.cell(1,3,'Priority H').font = Font(bold=True)
for ri, (cat, count) in enumerate(sorted(cats.items()), 2):
    ws2.cell(ri,1,cat)
    ws2.cell(ri,2,count)
    h_count = sum(1 for c,q,a,p,pr in QA if c==cat and pr=='H')
    ws2.cell(ri,3,h_count)
ws2.cell(len(cats)+3,1,'TOTAL').font = Font(bold=True)
ws2.cell(len(cats)+3,2,len(QA)).font = Font(bold=True)
ws2.column_dimensions['A'].width = 30
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 15

out_path = r'C:\Users\Willim Ho\Desktop\網站資料\Reylong_Chatbot_QA.xlsx'
wb.save(out_path)
print(f'輸出完成: {out_path}')
print(f'共 {len(QA)} 筆 Q&A')
for cat, count in sorted(cats.items()):
    h = sum(1 for c,q,a,p,pr in QA if c==cat and pr=='H')
    print(f'  {cat}: {count} 筆（H優先: {h}筆）')
