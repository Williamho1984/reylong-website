import openpyxl, sys, io
from openpyxl.styles import PatternFill

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\dev\marketing strategy\reylong-leads\output\leads.xlsx')
ws = wb['買主名單']
headers = [cell.value for cell in ws[1]]

new_leads = [
    # === Tier 1: 三封機直接對口 (三方袋/チャック袋/スタンドパック専門) ===
    {
        '公司名稱': '株式会社生産日本社（セイニチ）- Ziplock Bag Manufacturer',
        '國家': 'JP',
        '產品描述': 'Seinichi is the world\'s first inventor of integrated zipper bags. Products include Unipak (zipper PE bags), Lamizip (zipper laminated bags), Lamigrip (CPP sealant), Slider bags, Stand-up pouches, Spout pouches. Largest zipper bag brand in Japan. Direct end-user of JL-L-2TZP600 zipper and three-side seal functions.',
        '官方網址': 'seinichi.co.jp',
        '聯絡頁面': None,
        '來源平台': 'JPI 日本包装技術協会 / 業界調查',
        '優先級': '⭐ 高優先',
        '標記關鍵字': 'zipper bag,three-side seal,laminated,doypack,high volume',
        '開發': None,
        '草稿狀態': '待生成',
        '最後聯繫日期': None,
    },
    {
        '公司名稱': '大伸産業株式会社 - Laminated Film Bag Specialist',
        '國家': 'JP',
        '產品描述': 'Daishin Sangyo specializes in laminated film bag manufacturing with 50+ years experience. Products: 三方袋 (three-side seal bags), チャック付き三方袋 (zipper three-side), スタンドパック (Doypack), チャック付きスタンドパック (zipper Doypack). Applications: retort food, snacks, tea, medical. Perfect fit for JL-L-2TZP600.',
        '官方網址': 'daishinsangyo.com',
        '聯絡頁面': None,
        '來源平台': '業界調查 / Metoree',
        '優先級': '⭐ 高優先',
        '標記關鍵字': 'three-side seal,zipper,doypack,laminated film,food packaging',
        '開發': None,
        '草稿狀態': '待生成',
        '最後聯繫日期': None,
    },
    {
        '公司名稱': '細川洋行 - Pioneer Flexible Packaging Manufacturer',
        '國家': 'JP',
        '產品描述': 'Hosokawa Yoko is a pioneer (軟包装業界の先駆者) in the Japanese flexible packaging industry. Produces high-quality food packaging including pouches for snacks, retort foods, and frozen foods. Delivers flexible bags with easy-open and re-closable features. Strong alignment with JL-L-2TZP600 heat-seal laminated film bags.',
        '官方網址': 'hosokawa-yoko.co.jp',
        '聯絡頁面': None,
        '來源平台': 'JPI 日本包装技術協会 / 業界調查',
        '優先級': '⭐ 高優先',
        '標記關鍵字': 'flexible packaging,food,retort,pouch,three-side seal',
        '開發': None,
        '草稿狀態': '待生成',
        '最後聯繫日期': None,
    },
    {
        '公司名稱': 'クリロン化成株式会社 - Composite Film & Bag Manufacturer',
        '國家': 'JP',
        '產品描述': 'Kurilon Kasei is a composite film (複合フィルム) manufacturer making a wide range of laminated bags for food, medical, and industrial use. Known for Kyobi-jin laminated vacuum bags and standard-size flexible bags. Uses heat-seal laminated films matching JL-L-2TZP600 material specifications.',
        '官方網址': 'kurilon.co.jp',
        '聯絡頁面': None,
        '來源平台': 'JPI 日本包装技術協会',
        '優先級': '⭐ 高優先',
        '標記關鍵字': 'laminated film,flexible bag,heat-seal,food,medical',
        '開發': None,
        '草稿狀態': '待生成',
        '最後聯繫日期': None,
    },
    {
        '公司名稱': '日本マタイ株式会社 - Bag Making & Resin Processing',
        '國家': 'JP',
        '產品描述': 'Nippon Matai uses bag making (製袋加工) and resin processing technology to produce medical/food packaging materials, fertilizer bags, and container bags. Diverse product range across industries. Has in-house bag making equipment and strong alignment with flexible pouch and bag production needs.',
        '官方網址': 'matai.co.jp',
        '聯絡頁面': None,
        '來源平台': 'JPI 日本包装技術協会',
        '優先級': '⭐ 高優先',
        '標記關鍵字': 'bag making,flexible packaging,food,medical,laminated',
        '開發': None,
        '草稿狀態': '待生成',
        '最後聯繫日期': None,
    },
    # === Tier 2: 関連軟包装メーカー (flexible packaging converters) ===
    {
        '公司名稱': 'アイセロ株式会社 - Specialty Film & Flexible Packaging',
        '國家': 'JP',
        '產品描述': 'Aicello Corporation produces specialty packaging films and flexible bags for industrial, food, and pharmaceutical sectors. Member of Japan Packaging Institute (JPI). Manufactures heat-sealable films and flexible pouches for diverse applications.',
        '官方網址': 'aicello.co.jp',
        '聯絡頁面': None,
        '來源平台': 'JPI 日本包装技術協会',
        '優先級': '⭐ 高優先',
        '標記關鍵字': 'specialty film,flexible packaging,heat-seal,pharmaceutical',
        '開發': None,
        '草稿狀態': '待生成',
        '最後聯繫日期': None,
    },
    {
        '公司名稱': 'キョーラク株式会社 - Flexible Packaging Manufacturer',
        '國家': 'JP',
        '產品描述': 'Kyoraku Co. is a flexible packaging manufacturer producing a wide range of packaging solutions including soft pouches, flexible containers, and packaging films. JPI member company with broad flexible packaging expertise relevant to JL-L-2TZP600.',
        '官方網址': 'krk.co.jp',
        '聯絡頁面': None,
        '來源平台': 'JPI 日本包装技術協会',
        '優先級': '⭐ 高優先',
        '標記關鍵字': 'flexible packaging,pouch,film,soft container',
        '開發': None,
        '草稿狀態': '待生成',
        '最後聯繫日期': None,
    },
    {
        '公司名稱': '中本パックス株式会社 - Flexible Packaging Printing',
        '國家': 'JP',
        '產品描述': 'Nakamoto Packs specializes in flexible packaging printing and conversion. Produces printed flexible bags and pouches for food, cosmetics, and industrial use. JPI member. Has gravure printing and bag-making capability aligned with JL-L-2TZP600 operations.',
        '官方網址': 'npacks.co.jp',
        '聯絡頁面': None,
        '來源平台': 'JPI 日本包装技術協会',
        '優先級': '⭐ 高優先',
        '標記關鍵字': 'flexible packaging,printing,gravure,pouch,laminated',
        '開發': None,
        '草稿狀態': '待生成',
        '最後聯繫日期': None,
    },
    {
        '公司名稱': '王子製袋株式会社 - Bag Manufacturing (Oji Group)',
        '國家': 'JP',
        '產品描述': 'Oji Pack Co. is the bag manufacturing arm of the Oji Paper Group, one of Japan\'s largest paper companies. Produces various types of packaging bags including flexible and composite bags. Strong manufacturing scale and diverse customer base.',
        '官方網址': 'ojipack.co.jp',
        '聯絡頁面': None,
        '來源平台': 'JPI 日本包装技術協会',
        '優先級': '⭐ 高優先',
        '標記關鍵字': 'bag manufacturing,flexible,composite,Oji Paper Group',
        '開發': None,
        '草稿狀態': '待生成',
        '最後聯繫日期': None,
    },
    {
        '公司名稱': '大石産業株式会社 - Film & Heavy Packaging Bags',
        '國家': 'JP',
        '產品描述': 'Oishi Sangyo is a comprehensive packaging materials manufacturer with four business pillars: pulp mold, film, heavy packaging bags, and corrugated cardboard. Flexible film bag division is relevant for JL-L-2TZP600. Mid-to-large scale manufacturer.',
        '官方網址': 'osk.co.jp',
        '聯絡頁面': None,
        '來源平台': 'JPI 日本包装技術協会',
        '優先級': '⭐ 高優先',
        '標記關鍵字': 'film,packaging bag,flexible,heavy packaging',
        '開發': None,
        '草稿狀態': '待生成',
        '最後聯繫日期': None,
    },
    {
        '公司名稱': '進洋株式会社 - Flexible Packaging Manufacturer',
        '國家': 'JP',
        '產品描述': 'Shinyo Pack is a Japanese flexible packaging manufacturer producing printed laminated bags, pouches, and flexible packaging materials for food and consumer goods. JPI member with strong converter capabilities aligned with JL-L-2TZP600.',
        '官方網址': 'shinyo-pack.co.jp',
        '聯絡頁面': None,
        '來源平台': 'JPI 日本包装技術協会',
        '優先級': '⭐ 高優先',
        '標記關鍵字': 'flexible packaging,laminated,printed bag,food',
        '開發': None,
        '草稿狀態': '待生成',
        '最後聯繫日期': None,
    },
    {
        '公司名稱': 'カナエ株式会社 - Flexible Packaging',
        '國家': 'JP',
        '產品描述': 'Kanae Co. is a Japanese flexible packaging company producing various packaging materials including kraft bags and barrier ziplock bags. Member of Japan Packaging Institute. Produces flexible bags for food, pharmaceutical, and industrial applications.',
        '官方網址': 'kk-kanae.jp',
        '聯絡頁面': None,
        '來源平台': 'JPI 日本包装技術協会',
        '優先級': 'normal',
        '標記關鍵字': 'flexible packaging,zipper,barrier,food,pharmaceutical',
        '開發': None,
        '草稿狀態': '待生成',
        '最後聯繫日期': None,
    },
    {
        '公司名稱': '昭和パックス株式会社 - Packaging Manufacturer',
        '國家': 'JP',
        '產品描述': 'Showa Paxxs Corporation (est. 1935, Tokyo) manufactures industrial packaging including kraft heavy bags, PE bags, and mid-size bags for rice, pet food, and chemicals. JPI member. Has flexible bag production capability potentially upgradeable with JL-L-2TZP600.',
        '官方網址': 'showa-paxxs.co.jp',
        '聯絡頁面': None,
        '來源平台': 'JPI 日本包装技術協会',
        '優先級': 'normal',
        '標記關鍵字': 'packaging,kraft bag,PE bag,industrial',
        '開發': None,
        '草稿狀態': '待生成',
        '最後聯繫日期': None,
    },
    {
        '公司名稱': 'フタムラ化学株式会社 - Packaging Film Manufacturer',
        '國家': 'JP',
        '產品描述': 'Futamura Chemical is a packaging film manufacturer producing OPP, cellophane, and other flexible films used in food packaging. JPI member. Supplies heat-sealable films compatible with JL-L-2TZP600 material specs (30-180 um heat-seal laminated film).',
        '官方網址': 'futamura.co.jp',
        '聯絡頁面': None,
        '來源平台': 'JPI 日本包装技術協会',
        '優先級': 'normal',
        '標記關鍵字': 'OPP film,cellophane,heat-seal film,flexible packaging',
        '開發': None,
        '草稿狀態': '待生成',
        '最後聯繫日期': None,
    },
]

green_fill = PatternFill('solid', fgColor='E2EFDA')
yellow_fill = PatternFill('solid', fgColor='FFF2CC')

for lead in new_leads:
    row_values = [lead.get(h) for h in headers]
    ws.append(row_values)
    fill = green_fill if '高優先' in str(lead.get('優先級', '')) else yellow_fill
    for col in range(1, len(headers) + 1):
        ws.cell(row=ws.max_row, column=col).fill = fill

wb.save(r'C:\dev\marketing strategy\reylong-leads\output\leads.xlsx')
high = sum(1 for l in new_leads if '高優先' in str(l.get('優先級', '')))
print(f'新增 {len(new_leads)} 筆日本廠商，現在共 {ws.max_row - 1} 筆')
print(f'  高優先: {high} 筆（綠色）')
print(f'  一般: {len(new_leads) - high} 筆（黃色）')
