import openpyxl, sys, io
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb_src = openpyxl.load_workbook(r'C:\dev\marketing strategy\reylong-leads\output\leads.xlsx')
ws_src = wb_src['買主名單']
headers = [cell.value for cell in ws_src[1]]

exclude_domains_jp = [
    'jstories.media','credenceresearch.com','ipros.jp','ipros.com','mordorintelligence.com',
    'packagingstrategies.com','japanupclose.web-japan.org','patents.google.com',
    'packing-machine.machine-machinery.com','tsunagujapan.com','cnlipack.com',
    'yiruixingpackaging.com','vanik.com','longdapac.com','masarmedical.com.qa',
    'tsukamototeisou.jp','plasticstoday.com','naneikyo.com','amo-pack.com',
    'daishowasiko.com','packweb.biz',
]
exclude_domains_us = [
    'ensun.io','kokoquest.com','productmkr.com','baijiabags.com','herrains.com',
    'hezcypak.com','tradekey.com','plasticstoday.com','bagsupply.com',
    'kanplas.com','tgoldkamp.com','syntheticgrassstore.com','laundrybagsonline.com',
]
mfr_kw = [
    'manufacturer','製造','メーカー','converter','packaging company','produces',
    'bag making','flexible packaging','three-side seal','zipper','doypack','pouch',
    'poly bag','plastic bag','printing','film','laminated','heat-seal','gravure',
]
machine_fit_kw = [
    'three-side','zipper','doypack','stand-up','pouch','side seal','slider',
    'flexible packaging','laminated film','heat seal','printed bag','gravure',
    'three side','zipper bag','flexible','三封','ジッパー','スタンド','ラミネート',
]

us_rows, jp_rows = [], []
seen_jp_domains = set()

for row in ws_src.iter_rows(min_row=2, values_only=True):
    name, country, desc, domain, contact, source, priority, keywords, dev, draft, last = (list(row)+[None]*11)[:11]
    if not name:
        continue
    ct = str(country or '').upper()
    prio = str(priority or '')
    domain_l = str(domain or '').lower()
    all_text = ' '.join(str(v or '').lower() for v in [name, desc, domain, keywords])

    if ct == 'US' and '高優先' in prio:
        if not any(d in domain_l for d in exclude_domains_us):
            has_fit = any(m in all_text for m in machine_fit_kw)
            us_rows.append((list(row), has_fit))

    if ct == 'JP':
        if any(d in domain_l for d in exclude_domains_jp):
            continue
        # 只保留高優先，或來自 JPI / 業界調查的新增廠商
        is_curated = any(s in str(source or '').lower() for s in ['jpi', '業界調查', 'metoree', 'dee machine'])
        is_high_prio = '高優先' in prio
        if not (is_curated or is_high_prio):
            continue
        has_mfr = any(m in all_text for m in mfr_kw)
        if not has_mfr:
            continue
        # 去重：同一 domain 只取一筆
        if domain_l in seen_jp_domains:
            continue
        seen_jp_domains.add(domain_l)
        has_fit = any(m in all_text for m in machine_fit_kw)
        jp_rows.append((list(row), has_fit))

# ─── Build output workbook ───
wb_out = openpyxl.Workbook()

# Styles
header_fill   = PatternFill('solid', fgColor='1F4E79')
header_font   = Font(bold=True, color='FFFFFF', size=11)
fit_fill      = PatternFill('solid', fgColor='E2EFDA')   # green  = 三封機對口
general_fill  = PatternFill('solid', fgColor='DDEBF7')   # blue   = 一般製造商
alt_fill      = PatternFill('solid', fgColor='F2F2F2')

col_widths = [45, 8, 55, 28, 25, 30, 14, 30, 10, 12, 16]

def write_sheet(wb, title, rows_with_flag, country_label):
    ws = wb.create_sheet(title)
    # Header
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 22

    fit_count = 0
    for ri, (row, has_fit) in enumerate(rows_with_flag, 2):
        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.alignment = Alignment(wrap_text=True, vertical='top')
            c.fill = fit_fill if has_fit else general_fill
        if has_fit:
            fit_count += 1
        ws.row_dimensions[ri].height = 55

    # Add a tag column after last column
    tag_col = len(headers) + 1
    ws.cell(row=1, column=tag_col, value='三封機適配').fill = header_fill
    ws.cell(row=1, column=tag_col).font = header_font
    for ri, (row, has_fit) in enumerate(rows_with_flag, 2):
        c = ws.cell(row=ri, column=tag_col, value='✓ 對口' if has_fit else '—')
        c.alignment = Alignment(horizontal='center', vertical='top')

    # Column widths
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.column_dimensions[get_column_letter(tag_col)].width = 12
    ws.freeze_panes = 'A2'

    return len(rows_with_flag), fit_count

# Remove default sheet
wb_out.remove(wb_out.active)

us_total, us_fit = write_sheet(wb_out, '美國高優先名單', us_rows, 'US')
jp_total, jp_fit = write_sheet(wb_out, '日本製造商名單', jp_rows, 'JP')

# Summary sheet
ws_sum = wb_out.create_sheet('篩選摘要', 0)
ws_sum.column_dimensions['A'].width = 30
ws_sum.column_dimensions['B'].width = 20

summary = [
    ('篩選摘要', ''),
    ('', ''),
    ('美國高優先廠商', f'{us_total} 家'),
    ('  其中三封機對口', f'{us_fit} 家'),
    ('', ''),
    ('日本真實製造商', f'{jp_total} 家'),
    ('  其中三封機對口', f'{jp_fit} 家'),
    ('', ''),
    ('篩選說明', ''),
    ('美國', '高優先 + 排除目錄/聚合網站'),
    ('日本', '高優先 或 JPI/業界調查來源 + 真實製造商 + 去重'),
    ('三封機對口', '含 three-side/zipper/doypack/flexible packaging 等關鍵字'),
    ('機器型號', 'JL-L-2TZP600 多功能三封/夾鏈/Doypack 製袋機'),
    ('適用材料', 'Heat-seal laminated film 30–180 μm'),
]
for ri, (k, v) in enumerate(summary, 1):
    ws_sum.cell(row=ri, column=1, value=k).font = Font(bold=True if ri in [1,9] else False)
    ws_sum.cell(row=ri, column=2, value=v)

out_path = r'C:\dev\marketing strategy\reylong-leads\output\filtered_leads_JL-L-2TZP600.xlsx'
wb_out.save(out_path)
print(f'輸出完成: {out_path}')
print(f'  美國高優先: {us_total} 家（三封機對口 {us_fit} 家）')
print(f'  日本製造商: {jp_total} 家（三封機對口 {jp_fit} 家）')
